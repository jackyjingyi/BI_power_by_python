<<<<<<< HEAD
import pandas as pd
import openpyxl
from Lab_audit import get_input
from Lab_audit.get_Tracker_ifno import Extracker


location = '/Users/yanjingy/Documents/work/on_going_project/Lab_audit/'
file_names = ['Lab Audit Tracker.xlsx','lab Portal Audit WBR.xlsx','week_input.txt']

# if start - end date not given , use initial_date instead
initial_date = '2019-02-18'
period = 7

# collect info from input file
info = get_input.Pget(location + file_names[2])
info.get_lines()
info_store = [info.get_week(),info.get_start(),info.get_end()]
# get info 
week_number = info_store[0]
start_date = info_store[1]
end_date = info_store[2]
# collect useful raw data
tracker = Extracker(location+file_names[0])
tracker_overall_rating = tracker.overall_rating('Lab_group','Overall Rating')#,start_date = start_date,end_date = end_date)
tracker_tat = tracker.count_tat('Lab_group')#,start_date =start_date,end_date = end_date)
tracker_test_report = tracker.count_test_report('Lab_group')#,start_date =start_date,end_date = end_date)


# current target location
wb = openpyxl.load_workbook(filename=location+file_names[1])
ws = wb['Sheet3']


for i in range(1, len(ws['B'])):
    if ws['B'+str(i)].value is not None:
        temp_lab = ws['B'+str(i)].value
        print(temp_lab)
        ws['D'+str(i)] = tracker_test_report[temp_lab]
        temp_rating = ws['C'+str(i+1)].value
        ws['D'+str(i+1)] = '{:.2%}'.format(tracker_overall_rating[temp_lab,'Fail']/sum(tracker_overall_rating[temp_lab]))
        ws['D'+str(i+2)] = round(tracker_tat[temp_lab],2)
wb.save(location+file_names[1])
data_rows = []

for row in ws['B7':'G18']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# Transform into dataframe

df1 = pd.DataFrame(data_rows)
print(df1)
print(type(df1.iloc[1,1]))





=======
import pandas as pd
import openpyxl
from Lab_audit import get_input
from Lab_audit.get_Tracker_ifno import Extracker


location = '/Users/yanjingy/Documents/work/on_going_project/Lab_audit/'
file_names = ['Lab Audit Tracker.xlsx','lab Portal Audit WBR.xlsx','week_input.txt']

# if start - end date not given , use initial_date instead
initial_date = '2019-02-18'
period = 7

# collect info from input file
info = get_input.Pget(location + file_names[2])
info.get_lines()
info_store = [info.get_week(),info.get_start(),info.get_end()]
# get info 
week_number = info_store[0]
start_date = info_store[1]
end_date = info_store[2]
# collect useful raw data
tracker = Extracker(location+file_names[0])
tracker_overall_rating = tracker.overall_rating('Lab_group','Overall Rating')#,start_date = start_date,end_date = end_date)
tracker_tat = tracker.count_tat('Lab_group')#,start_date =start_date,end_date = end_date)
tracker_test_report = tracker.count_test_report('Lab_group')#,start_date =start_date,end_date = end_date)


# current target location
wb = openpyxl.load_workbook(filename=location+file_names[1])
ws = wb['Sheet3']


for i in range(1, len(ws['B'])):
    if ws['B'+str(i)].value is not None:
        temp_lab = ws['B'+str(i)].value
        print(temp_lab)
        ws['D'+str(i)] = tracker_test_report[temp_lab]
        temp_rating = ws['C'+str(i+1)].value
        ws['D'+str(i+1)] = '{:.2%}'.format(tracker_overall_rating[temp_lab,'Fail']/sum(tracker_overall_rating[temp_lab]))
        ws['D'+str(i+2)] = round(tracker_tat[temp_lab],2)
wb.save(location+file_names[1])
data_rows = []

for row in ws['B7':'G18']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

# Transform into dataframe

df1 = pd.DataFrame(data_rows)
print(df1)
print(type(df1.iloc[1,1]))





>>>>>>> adddbee2a85b4786c4c9a1c91ecccc5eb68a0bbf

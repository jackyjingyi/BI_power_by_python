
## this script actually do a sumifs job in excel
import openpyxl
import string
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.utils import coordinate_from_string,column_index_from_string
# excel location

# assume number greater than 0

loc = '/you file location'
wb = openpyxl.load_workbook(loc)

Tracker = wb['Tracker']
spec = wb['Spec Level']
a_list = []
ano_list = []
# get the header column names
for cell in spec[1]:
    a_list.append(cell.value)
for cell in Tracker[1]:
    ano_list.append(cell.value)
# zip it as a dictionary {column name : index}
col_indices = {n:cell.value for n, cell in enumerate(spec[1]) if cell.value in a_list}
reverse_col_index_spec = dict(zip(col_indices.values(),col_indices.keys()))
col_indices_tracker = {n:cell.value for n, cell in enumerate(Tracker[1]) if cell.value in ano_list}
reverse_col_index_tracker = dict(zip(col_indices_tracker.values(),col_indices_tracker.keys()))
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
#                                                             #
#  reverse_col_index_spec : spec column name dictionary       #
#  reverse_col_index_tracker : tracker column name dictionary #
#                                                             #
# # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

# all MFR part, to be reuse, this function return specific column of a given worksheet
def extractCol(sheet,identifier):
    identifer_list = list(sheet[get_column_letter(reverse_col_index_spec[identifier]+1)])
    return identifer_list

# remove duplicates from
the_set =set(
    [item.value for item in list(
    spec[get_column_letter(reverse_col_index_spec["Issue found with the test/test method (reason for fail)"]+1)])
    ])

class Node:
    def __init__(self,key,value):
        self.key = key
        self.value = value

def count(sheet,identifier,col1,col2):    # col1&col2 is where restriction words located we count for
    import copy
    iden_list = list(set([value.value for value in extractCol(sheet,identifier)]))  # remove duplicates of dictionary's keys
    idenNumber = reverse_col_index_spec[identifier]                                 # the identifier column
    issueNumber = reverse_col_index_spec[col1]                                      # the issue number column
    safetyNumber = reverse_col_index_spec[col2]                                     # the safety number column

    # create a dictionary {identifier: {issue1 :{safety:0,regulatory:0},issue2:{etc}}}
    # # # # # # # # # # # # # # # # # # # # # # # # # #
    safety_regulatory = {'Safety' :0 , 'Regulatory' :0}
    aux_list_dict = {}
    for item in list(the_set):
        aux_list_dict[item] = copy.deepcopy(safety_regulatory)
    
    the_info_list = []
    dict_store = [None] * len(iden_list)
    for i in range(len(dict_store)):
        dict_store[i]=copy.deepcopy(aux_list_dict)

    check_dict = dict(zip(iden_list,dict_store))
    # # # # # # # # # # # # # # # # # # # # # # # # # # 
    
    # update count number
    for row in sheet.iter_rows():
        currentvalue = row[idenNumber].value
        current_issue = row[issueNumber].value
        current_safety = row[safetyNumber].value
        if currentvalue is not None and current_issue is not None and (current_safety == "Safety" or current_safety =="Regulatory"):
            check_dict[currentvalue][current_issue][current_safety]+=1
            print(currentvalue,current_issue,current_safety,check_dict[currentvalue][current_issue][current_safety])

    return the_info_list


a = count(spec,"MFR Part","Issue found with the test/test method (reason for fail)","Safety & Regulatory")
print(a)
